import customtkinter
import requests
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import Workbook
from urllib.parse import urlparse, urljoin


def is_valid_url(url, base_url):
	"""Проверяет, является ли URL действительным и относится ли к базовому URL."""
	parsed = urlparse(url)
	return bool(parsed.netloc) and parsed.netloc == urlparse(base_url).netloc


def get_page_data(url):
	"""Получает данные о заголовке, описании и тегах h1-h6 страницы."""
	response = requests.get(url)
	response.raise_for_status()
	soup = BeautifulSoup(response.text, 'html.parser')

	title = soup.title.string if soup.title else ''
	description = soup.find('meta', attrs={'name': 'description'})[
		'content'] if soup.find('meta', attrs={'name': 'description'}) else ''
	h1 = soup.find('h1').text if soup.find('h1') else ''
	h2 = soup.find('h2').text if soup.find('h2') else ''
	h3 = soup.find('h3').text if soup.find('h3') else ''
	h4 = soup.find('h4').text if soup.find('h4') else ''
	h5 = soup.find('h5').text if soup.find('h5') else ''
	h6 = soup.find('h6').text if soup.find('h6') else ''

	return title, description, h1, h2, h3, h4, h5, h6


def crawl_website():
	url = entry.get()

	try:
		response = requests.get(url)
		response.raise_for_status()
		soup = BeautifulSoup(response.text, 'html.parser')
		base_url = urlparse(url).scheme + "://" + urlparse(url).netloc

		links = [urljoin(base_url, link.get('href')) for link in
		         soup.find_all('a')]
		pages = list(filter(lambda link: is_valid_url(link, base_url), links))

		total_pages = len(pages)

		if total_pages == 0:
			messagebox.showinfo("Информация",
			                    "На веб-сайте не найдено страниц.")
			return

		save_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
		if save_path:
			workbook = Workbook()
			sheet = workbook.active

			# Заголовки столбцов
			sheet['A1'] = 'URL'
			sheet['B1'] = 'Title'
			sheet['C1'] = 'Description'
			sheet['D1'] = 'H1'
			sheet['E1'] = 'H2'
			sheet['F1'] = 'H3'
			sheet['G1'] = 'H4'
			sheet['H1'] = 'H5'
			sheet['I1'] = 'H6'

			for index, page in enumerate(pages, start=2):
				title, description, h1, h2, h3, h4, h5, h6 = get_page_data(
					page)

				sheet.cell(row=index, column=1).value = page
				sheet.cell(row=index, column=2).value = title
				sheet.cell(row=index, column=3).value = description
				sheet.cell(row=index, column=4).value = h1
				sheet.cell(row=index, column=5).value = h2
				sheet.cell(row=index, column=6).value = h3
				sheet.cell(row=index, column=7).value = h4
				sheet.cell(row=index, column=8).value = h5
				sheet.cell(row=index, column=9).value = h6

				# Обновление счетчика прогресса
				progress_label.configure(
					text=f"Просканировано {index - 1} из {total_pages}")
				root.update()

			workbook.save(save_path)

			messagebox.showinfo("Готово!",
			                    f"Страницы успешно сохранены! Сохранено: {total_pages}")

			# Отображение кнопки "Open File" для открытия сохраненного файла
			def open_file():
				import os
				os.startfile(save_path)

			button_open = tk.Button(root, text="Open File", command=open_file)
			button_open.pack(pady=10)
		else:
			messagebox.showinfo("Информация", "Сохранение файла отменено.")

		# Очистка метки и поля ввода
		entry.delete(0, tk.END)
		progress_label.config(text="")

	except requests.exceptions.RequestException as e:
		messagebox.showerror("Ошибка", f"Произошла ошибка: {e}")


# Создание главного окна
root = customtkinter.CTk()
root.title("OneScanner")
customtkinter.set_default_color_theme("green")
customtkinter.set_appearance_mode("dark")
# Размер окна
window_width = 330
window_height = 180

# Определение положения окна на экране
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = int((screen_width / 2) - (window_width / 2))
y = int((screen_height / 2) - (window_height / 2))
root.geometry(f"{window_width}x{window_height}+{x}+{y}")

# Создание метки и поля ввода для URL веб-сайта
# label = tk.Label(root, text="Введите URL веб-сайта:")
# label.pack()
entry = customtkinter.CTkEntry(root, placeholder_text="Введите URL веб-сайта:", width=250)
entry.pack(pady=20)

# Создание кнопки для запуска процесса сканирования
button = customtkinter.CTkButton(root, text="Сканировать", command=crawl_website)
button.pack(pady=10)

# Метка для отображения прогресса
progress_label = customtkinter.CTkLabel(root, text="")
progress_label.pack(pady=10)

# Запуск цикла обработки событий Tkinter
root.mainloop()
